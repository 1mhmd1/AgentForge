"""
Safe Code Injector
Injects content into Python agent templates using serialization layer
Prevents syntax errors from unsafe string embedding
"""

import base64
import csv
import io
import json
from typing import Dict, Any
from services.code_serializer import serialize_for_domain, CodeSerializer
from services.code_sanitizer import pre_syntax_check


def _xml_element_to_dict_eager(el):
    node = {"tag": el.tag}
    if el.attrib:
        node["attrs"] = dict(el.attrib)
    children = list(el)
    if children:
        node["children"] = [_xml_element_to_dict_eager(c) for c in children]
    elif el.text and el.text.strip():
        node["text"] = el.text.strip()
    return node


def _eager_transform(mimetype: str, b64: str) -> str:
    """
    Build-time twin of parse_input(): parses the uploaded bytes and returns a
    JSON string. Used to pre-populate TRANSFORMED_OUTPUT in the generated
    agent so the frontend can offer a download immediately on `success`.
    """
    raw = base64.b64decode(b64)
    m = (mimetype or "").lower()
    if "csv" in m and "tab" not in m:
        text = raw.decode("utf-8")
        parsed = list(csv.DictReader(io.StringIO(text)))
    elif "tab-separated" in m or m.endswith("tsv"):
        text = raw.decode("utf-8")
        parsed = list(csv.DictReader(io.StringIO(text), delimiter="\t"))
    elif "ndjson" in m or "jsonl" in m:
        text = raw.decode("utf-8")
        parsed = [json.loads(line) for line in text.splitlines() if line.strip()]
    elif "json" in m:
        parsed = json.loads(raw.decode("utf-8"))
    elif "openxmlformats" in m or m.endswith("xlsx"):
        # MUST come before the generic "xml" branch -- the xlsx mimetype is
        # `application/vnd.openxmlformats-...` which contains the substring
        # "xml" and would otherwise be mis-routed to the XML parser.
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parsed = []
        else:
            header = [str(h) if h is not None else "" for h in rows[0]]
            parsed = [
                {header[i]: r[i] for i in range(min(len(header), len(r)))}
                for r in rows[1:]
            ]
    elif "xml" in m:
        from xml.etree import ElementTree as ET
        parsed = _xml_element_to_dict_eager(ET.fromstring(raw.decode("utf-8")))
    else:
        parsed = {"text": raw.decode("utf-8", errors="replace")}
    return json.dumps(parsed, indent=2, ensure_ascii=False, default=str)


class SafeCodeInjector:
    """Safely inject content into Python agent code"""

    AGENT_TEMPLATE = '''import os
import json
from datetime import datetime, timezone

# ── Injected Metadata ──────────────────────────────────────
RUN_ID = "{run_id}"
OUTPUT_DIR = "{output_dir}"
DOMAIN = "{domain}"
GOAL = "{goal}"
COMPLEXITY = "{complexity}"
SUCCESS_CRITERIA = "{success_criteria}"
STEPS = {steps}
TOOLS = {tools}
INPUTS = []
OUTPUTS = []

{content_section}

# ── Sub-Agent Functions ────────────────────────────────────
{agent_functions}

# ── Main Execution ─────────────────────────────────────────
def main():
    """Execute agent pipeline"""
    print(f"Starting {{DOMAIN}} pipeline: {{GOAL}}")

    try:
{main_body}

        # Save output
        save_output(final_result)
        return final_result

    except Exception as e:
        print(f"Execution error: {{str(e)}}")
        raise

def save_output(data):
    """Save final output to file"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    extensions = {{
        "website_builder": "html",
        "document": "md",
        "web_research": "txt",
        "data_transform": "json"
    }}
    ext = extensions.get(DOMAIN, "txt")

    filepath = f"{{OUTPUT_DIR}}/{{RUN_ID}}_{{DOMAIN}}.{{ext}}"

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(str(data))

    print(f"Saved to {{filepath}}")

if __name__ == "__main__":
    main()
'''

    @staticmethod
    def inject_safe(
        domain: str,
        goal: str,
        steps: list,
        tools: list,
        content: Dict[str, Any],
        agent_functions: str,
        main_body: str,
        run_id: str = "ui_generated",
        complexity: str = "simple",
        success_criteria: str = "Task completed"
    ) -> str:
        """
        Safely inject content into agent template

        Returns:
            Complete Python agent code
        """
        content_section = serialize_for_domain(domain, **content)

        serializer = CodeSerializer()
        steps_safe = json.dumps(steps)
        tools_safe = json.dumps(tools)

        code = SafeCodeInjector.AGENT_TEMPLATE.format(
            run_id=run_id,
            output_dir="output",
            domain=domain,
            goal=serializer.escape_for_python(goal),
            complexity=complexity,
            success_criteria=serializer.escape_for_python(success_criteria),
            steps=steps_safe,
            tools=tools_safe,
            content_section=content_section,
            agent_functions=agent_functions,
            main_body=main_body
        )

        return code

    @staticmethod
    def build_website_agent(
        goal: str,
        html: str,
        css: str = "",
        js: str = "",
        run_id: str = "ui_generated"
    ) -> str:
        """Build website_builder domain agent"""

        agent_functions = '''def step_1_generate_page():
    """Generate complete HTML page"""
    full_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{GOAL}</title>
    <style>
    {CSS_CONTENT}
    </style>
</head>
<body>
{HTML_CONTENT}
    <script>
    {JS_CONTENT}
    </script>
</body>
</html>"""
    return full_html
'''

        main_body = '''        # Step 1: Generate page
        result = step_1_generate_page()
        print(f"Generated HTML: {len(result)} chars")

        final_result = result'''

        return SafeCodeInjector.inject_safe(
            domain="website_builder",
            goal=goal,
            steps=["Generate HTML page"],
            tools=["generate", "code"],
            content={"html": html, "css": css, "js": js},
            agent_functions=agent_functions,
            main_body=main_body,
            run_id=run_id
        )

    @staticmethod
    def build_research_agent(
        goal: str,
        query: str,
        run_id: str = "ui_generated"
    ) -> str:
        """Build web_research domain agent"""

        agent_functions = '''def step_1_search():
    """Search for information"""
    import requests
    sources = []
    search_url = f"https://www.google.com/search?q={RESEARCH_CONTENT}"
    try:
        resp = requests.get(search_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        sources.append({"url": search_url, "status": resp.status_code, "length": len(resp.text)})
    except Exception as e:
        sources.append({"error": str(e)})
    return json.dumps(sources, indent=2)

def step_2_summarize(search_results):
    """Summarize findings into a report"""
    report = f"""# Research Report: {GOAL}

## Sources Found
{search_results}

## Summary
Research completed for: {RESEARCH_CONTENT}

Generated: {datetime.now(timezone.utc).isoformat()}
"""
    return report
'''

        main_body = '''        # Step 1: Search
        search_results = step_1_search()
        print(f"Search complete: {len(search_results)} chars")

        # Step 2: Summarize (using step 1 output)
        summary = step_2_summarize(search_results)
        print(f"Summary complete: {len(summary)} chars")

        final_result = summary'''

        return SafeCodeInjector.inject_safe(
            domain="web_research",
            goal=goal,
            steps=["Search for information", "Summarize findings"],
            tools=["search", "summarize"],
            content={"content": query},
            agent_functions=agent_functions,
            main_body=main_body,
            run_id=run_id
        )

    @staticmethod
    def build_document_agent(
        goal: str,
        topic: str,
        run_id: str = "ui_generated"
    ) -> str:
        """Build document domain agent"""

        agent_functions = '''def step_1_research():
    """Research topic and gather key points"""
    key_points = [
        f"Overview of {DOCUMENT_CONTENT}",
        f"Key aspects and considerations",
        f"Current trends and developments",
        f"Best practices and recommendations",
        f"Future outlook and implications",
    ]
    return key_points

def step_2_format_report(findings):
    """Format findings as a structured report"""
    sections = "\\n".join(f"- {point}" for point in findings)
    report = f"""# Technical Report: {GOAL}

## Executive Summary
This report covers the key aspects of: {DOCUMENT_CONTENT}

## Key Findings
{sections}

## Conclusion
Report generated on {datetime.now(timezone.utc).isoformat()}
"""
    return report
'''

        main_body = '''        # Step 1: Research
        findings = step_1_research()
        print(f"Research complete: {len(findings)} points")

        # Step 2: Format report (using step 1 output)
        report = step_2_format_report(findings)
        print(f"Report complete: {len(report)} chars")

        final_result = report'''

        return SafeCodeInjector.inject_safe(
            domain="document",
            goal=goal,
            steps=["Research topic", "Format report"],
            tools=["generate", "summarize"],
            content={"content": topic},
            agent_functions=agent_functions,
            main_body=main_body,
            run_id=run_id
        )

    @staticmethod
    def build_data_agent(
        goal: str,
        data: Any,
        run_id: str = "ui_generated",
        input_filename: str | None = None,
        input_mimetype: str | None = None,
        input_bytes_b64: str | None = None,
    ) -> str:
        """
        Build data_transform domain agent.

        Two modes:
          1) `input_bytes_b64` set: the user uploaded a file. The generated
             agent base64-decodes it, parses based on mimetype, and emits the
             parsed structure as JSON. Covers CSV/TSV/JSON/JSONL/XML/XLSX.
          2) Otherwise (legacy): fall back to the older sort-and-stringify
             behaviour driven by `data`.
        """
        if input_bytes_b64:
            return SafeCodeInjector._build_data_agent_with_upload(
                goal=goal,
                run_id=run_id,
                input_filename=input_filename or "uploaded_file",
                input_mimetype=input_mimetype or "application/octet-stream",
                input_bytes_b64=input_bytes_b64,
            )

        agent_functions = '''def step_1_parse():
    """Parse input data"""
    return DATA_PARSED

def step_2_transform(data):
    """Transform data into output format"""
    if isinstance(data, dict):
        result = {k: v for k, v in sorted(data.items())}
    elif isinstance(data, list):
        result = sorted(data, key=lambda x: str(x))
    else:
        result = data
    return json.dumps(result, indent=2, ensure_ascii=False)
'''

        main_body = '''        # Step 1: Parse input
        data = step_1_parse()
        print(f"Parsed data: {type(data).__name__}")

        # Step 2: Transform (using step 1 output)
        transformed = step_2_transform(data)
        print(f"Transform complete: {len(transformed)} chars")

        final_result = transformed'''

        return SafeCodeInjector.inject_safe(
            domain="data_transform",
            goal=goal,
            steps=["Parse input data", "Transform to output"],
            tools=["analyze", "code", "validate"],
            content={"content": data},
            agent_functions=agent_functions,
            main_body=main_body,
            run_id=run_id
        )

    @staticmethod
    def _build_data_agent_with_upload(
        goal: str,
        run_id: str,
        input_filename: str,
        input_mimetype: str,
        input_bytes_b64: str,
    ) -> str:
        """
        Generate a self-contained data_transform agent that bakes the uploaded
        file (base64) into the source and emits a JSON conversion. The result
        is also surfaced as TRANSFORMED_OUTPUT at module scope so the frontend
        can extract + offer download without re-running the agent.
        """
        from services.code_serializer import CodeSerializer

        serializer = CodeSerializer()
        goal_safe = serializer.escape_for_python(goal)

        # Trying to also compute the converted output at BUILD TIME so we can
        # embed it as a JSON-encoded module-level constant. The frontend reads
        # this with the same JsonStringVar reader used for HTML_CONTENT and
        # turns it into a Download button. Best-effort -- on parse failure we
        # ship a stub and let the validator's executed run produce the file.
        try:
            transformed_text = _eager_transform(input_mimetype, input_bytes_b64)
        except Exception:
            transformed_text = ""

        import json as _json
        transformed_literal = _json.dumps(transformed_text)
        input_filename_literal = _json.dumps(input_filename)
        input_mimetype_literal = _json.dumps(input_mimetype)
        input_b64_literal = _json.dumps(input_bytes_b64)
        goal_literal = _json.dumps(goal_safe)
        run_id_literal = _json.dumps(run_id)

        return f'''import os
import io
import csv
import json
import base64
from datetime import datetime, timezone

# Injected metadata
RUN_ID = {run_id_literal}
OUTPUT_DIR = "output"
DOMAIN = "data_transform"
GOAL = {goal_literal}
INPUT_FILENAME = {input_filename_literal}
INPUT_MIMETYPE = {input_mimetype_literal}
INPUT_BYTES_B64 = {input_b64_literal}
TRANSFORMED_OUTPUT = {transformed_literal}


def parse_input(raw_bytes, mimetype):
    """Parse the uploaded bytes into a Python structure based on mimetype."""
    m = (mimetype or "").lower()
    if "csv" in m and "tab" not in m:
        text = raw_bytes.decode("utf-8")
        return list(csv.DictReader(io.StringIO(text)))
    if "tab-separated" in m or m.endswith("tsv"):
        text = raw_bytes.decode("utf-8")
        return list(csv.DictReader(io.StringIO(text), delimiter="\\t"))
    if "ndjson" in m or "jsonl" in m:
        text = raw_bytes.decode("utf-8")
        return [json.loads(line) for line in text.splitlines() if line.strip()]
    if "json" in m:
        return json.loads(raw_bytes.decode("utf-8"))
    if "openxmlformats" in m or m.endswith("xlsx"):
        # Must come BEFORE the "xml" check because the xlsx mimetype
        # `application/vnd.openxmlformats-...` contains the substring "xml".
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h) if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            out.append({{header[i]: r[i] for i in range(min(len(header), len(r)))}})
        return out
    if "xml" in m:
        from xml.etree import ElementTree as ET
        root = ET.fromstring(raw_bytes.decode("utf-8"))
        return _xml_element_to_dict(root)
    # Fallback: best-effort utf-8 text wrapped in {{"text": ...}}
    return {{"text": raw_bytes.decode("utf-8", errors="replace")}}


def _xml_element_to_dict(el):
    """Compact XML -> dict converter that keeps attrs + children + text."""
    node = {{"tag": el.tag}}
    if el.attrib:
        node["attrs"] = dict(el.attrib)
    children = list(el)
    if children:
        node["children"] = [_xml_element_to_dict(c) for c in children]
    elif el.text and el.text.strip():
        node["text"] = el.text.strip()
    return node


def main():
    print(f"Converting {{INPUT_FILENAME}} ({{INPUT_MIMETYPE}}) for goal: {{GOAL}}")
    raw = base64.b64decode(INPUT_BYTES_B64)
    parsed = parse_input(raw, INPUT_MIMETYPE)
    output = json.dumps(parsed, indent=2, ensure_ascii=False, default=str)
    save_output(output)
    return output


def save_output(text):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, f"{{RUN_ID}}_data_transform.json")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"Saved to {{filepath}} ({{len(text)}} chars)")


if __name__ == "__main__":
    main()
'''

    @staticmethod
    def build_and_validate(domain: str, **kwargs) -> Dict[str, Any]:
        """
        Build agent and validate before returning

        Returns:
            {"code": str, "valid": bool, "warnings": list, "error": str|None}
        """
        builders = {
            "website_builder": SafeCodeInjector.build_website_agent,
            "web_research": SafeCodeInjector.build_research_agent,
            "document": SafeCodeInjector.build_document_agent,
            "data_transform": SafeCodeInjector.build_data_agent,
        }

        builder = builders.get(domain)
        if not builder:
            return {
                "code": None,
                "valid": False,
                "warnings": [],
                "error": f"Unknown domain: {domain}"
            }

        # Each builder accepts its own kwarg set. We pass through everything
        # the caller supplied; Python will surface a TypeError if an unknown
        # kwarg leaks through (intentional -- the call site is wrong).
        code = builder(**kwargs)

        sanitized_code, warnings, is_safe = pre_syntax_check(code)

        return {
            "code": sanitized_code,
            "valid": is_safe,
            "warnings": warnings,
            "error": None if is_safe else "Syntax validation failed"
        }
